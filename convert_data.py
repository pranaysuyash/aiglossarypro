#!/usr/bin/env python3
"""
AI/ML Glossary Data Converter
Converts between Excel, CSV, and JSON formats for use with different processors.
"""

import pandas as pd
import json
import openpyxl
import argparse
import sys
from pathlib import Path

def excel_to_csv(excel_file, csv_file):
    """Convert Excel to CSV format."""
    print(f"Converting {excel_file} to {csv_file}...")
    
    # Method 1: Using pandas (handles most cases)
    try:
        df = pd.read_excel(excel_file)
        df.to_csv(csv_file, index=False, encoding='utf-8')
        print(f"✅ Successfully converted using pandas: {len(df)} rows, {len(df.columns)} columns")
        return True
    except Exception as e:
        print(f"⚠️ Pandas method failed: {e}")
        
    # Method 2: Using openpyxl (preserves exact Excel data)
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Extract headers
        headers = [cell.value for cell in ws[1]]
        
        # Extract all rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append([cell if cell is not None else '' for cell in row])
        
        # Create DataFrame and save as CSV
        df = pd.DataFrame(rows, columns=headers)
        df.to_csv(csv_file, index=False, encoding='utf-8')
        print(f"✅ Successfully converted using openpyxl: {len(df)} rows, {len(df.columns)} columns")
        return True
    except Exception as e:
        print(f"❌ Openpyxl method failed: {e}")
        return False

def excel_to_json(excel_file, json_file):
    """Convert Excel to JSON format."""
    print(f"Converting {excel_file} to {json_file}...")
    
    # Method 1: Using pandas
    try:
        df = pd.read_excel(excel_file)
        df.to_json(json_file, orient='records', indent=2, force_ascii=False)
        print(f"✅ Successfully converted using pandas: {len(df)} records")
        return True
    except Exception as e:
        print(f"⚠️ Pandas method failed: {e}")
    
    # Method 2: Using openpyxl (preserves exact Excel data)
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Extract headers
        headers = [cell.value for cell in ws[1]]
        
        # Extract all rows and convert to list of dictionaries
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    record[headers[i]] = cell if cell is not None else ''
            records.append(record)
        
        # Save as JSON
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(records, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Successfully converted using openpyxl: {len(records)} records")
        return True
    except Exception as e:
        print(f"❌ Openpyxl method failed: {e}")
        return False

def csv_to_json(csv_file, json_file):
    """Convert CSV to JSON format."""
    print(f"Converting {csv_file} to {json_file}...")
    
    try:
        df = pd.read_csv(csv_file)
        df.to_json(json_file, orient='records', indent=2, force_ascii=False)
        print(f"✅ Successfully converted: {len(df)} records")
        return True
    except Exception as e:
        print(f"❌ Conversion failed: {e}")
        return False

def json_to_csv(json_file, csv_file):
    """Convert JSON to CSV format."""
    print(f"Converting {json_file} to {csv_file}...")
    
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        df = pd.DataFrame(data)
        df.to_csv(csv_file, index=False, encoding='utf-8')
        print(f"✅ Successfully converted: {len(df)} rows")
        return True
    except Exception as e:
        print(f"❌ Conversion failed: {e}")
        return False

def validate_file(filepath):
    """Validate that file exists and is readable."""
    if not Path(filepath).exists():
        print(f"❌ File not found: {filepath}")
        return False
    
    try:
        size = Path(filepath).stat().st_size
        print(f"📁 File found: {filepath} ({size:,} bytes)")
        return True
    except Exception as e:
        print(f"❌ Cannot access file: {e}")
        return False

def main():
    parser = argparse.ArgumentParser(description="Convert AI/ML Glossary data between formats")
    parser.add_argument('input_file', help='Input file path')
    parser.add_argument('output_file', help='Output file path')
    parser.add_argument('--validate', action='store_true', help='Validate conversion by checking row/column counts')
    
    args = parser.parse_args()
    
    input_file = args.input_file
    output_file = args.output_file
    
    # Validate input file
    if not validate_file(input_file):
        sys.exit(1)
    
    # Determine conversion type based on file extensions
    input_ext = Path(input_file).suffix.lower()
    output_ext = Path(output_file).suffix.lower()
    
    success = False
    
    if input_ext in ['.xlsx', '.xls'] and output_ext == '.csv':
        success = excel_to_csv(input_file, output_file)
    elif input_ext in ['.xlsx', '.xls'] and output_ext == '.json':
        success = excel_to_json(input_file, output_file)
    elif input_ext == '.csv' and output_ext == '.json':
        success = csv_to_json(input_file, output_file)
    elif input_ext == '.json' and output_ext == '.csv':
        success = json_to_csv(input_file, output_file)
    else:
        print(f"❌ Unsupported conversion: {input_ext} → {output_ext}")
        print("Supported conversions:")
        print("  Excel (.xlsx/.xls) → CSV (.csv)")
        print("  Excel (.xlsx/.xls) → JSON (.json)")
        print("  CSV (.csv) → JSON (.json)")
        print("  JSON (.json) → CSV (.csv)")
        sys.exit(1)
    
    if success:
        # Validate output file
        if validate_file(output_file):
            print(f"🎉 Conversion completed successfully!")
            
            if args.validate:
                print("\n🔍 Validation:")
                try:
                    if output_ext == '.csv':
                        df = pd.read_csv(output_file)
                        print(f"   CSV: {len(df)} rows × {len(df.columns)} columns")
                    elif output_ext == '.json':
                        with open(output_file, 'r') as f:
                            data = json.load(f)
                        print(f"   JSON: {len(data)} records")
                        if data:
                            print(f"   Fields: {len(data[0])} per record")
                except Exception as e:
                    print(f"   ⚠️ Validation error: {e}")
        else:
            print(f"❌ Output file was not created properly")
            sys.exit(1)
    else:
        print(f"❌ Conversion failed")
        sys.exit(1)

if __name__ == "__main__":
    main()